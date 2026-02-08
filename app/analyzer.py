"""
Auto-Assessment Analyzer

Two modes:
1. AI-powered (Claude API): Sends document text + each criterion to Claude for
   contextual analysis with proper understanding of SSBJ requirements.
2. Keyword fallback: Simple keyword matching when no API key is configured.

Text extraction supports PDF, DOCX, XLSX, CSV, TXT.
"""

import os
import csv
import hashlib
import json
import logging
import re
import time

from app.ssbj_criteria import SSBJ_CRITERIA, MATURITY_LEVELS

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Content-hash assessment cache
# ---------------------------------------------------------------------------

_assessment_cache = {}  # {hash: {"results": dict, "ts": float}}
_CACHE_TTL = 3600  # 1 hour


def _content_hash(text):
    """SHA256 prefix of document text — used as cache key."""
    return hashlib.sha256(text.encode("utf-8", errors="replace")).hexdigest()[:16]


def get_cached_results(combined_text):
    """Return cached assessment results or None."""
    h = _content_hash(combined_text)
    entry = _assessment_cache.get(h)
    if entry and time.time() - entry["ts"] < _CACHE_TTL:
        logger.info(f"Assessment cache hit ({h})")
        return entry["results"]
    if entry:
        del _assessment_cache[h]
    return None


def _cache_results(combined_text, results):
    """Store assessment results in cache."""
    h = _content_hash(combined_text)
    _assessment_cache[h] = {"results": results, "ts": time.time()}
    # Evict oldest if cache too large
    if len(_assessment_cache) > 20:
        oldest = min(_assessment_cache, key=lambda k: _assessment_cache[k]["ts"])
        del _assessment_cache[oldest]


def clear_cache():
    """Clear the assessment cache (useful after document changes)."""
    _assessment_cache.clear()


# ---------------------------------------------------------------------------
# Text extraction
# ---------------------------------------------------------------------------

def extract_text_from_file(filepath):
    """Extract text content from a file based on its extension.

    Returns (text, warning) tuple where warning is None on success,
    or a user-friendly message describing why extraction was partial/failed.
    For backward compatibility, callers that expect a plain string will still
    work — the result is truthy when text was extracted.
    """
    ext = filepath.rsplit(".", 1)[-1].lower() if "." in filepath else ""
    fname = os.path.basename(filepath)

    try:
        if ext == "pdf":
            return _extract_pdf(filepath)
        elif ext == "docx":
            return _extract_docx(filepath)
        elif ext in ("xlsx", "xls"):
            return _extract_xlsx(filepath)
        elif ext == "csv":
            return _extract_csv(filepath)
        elif ext == "txt":
            return _extract_txt(filepath)
        else:
            return ""
    except Exception as e:
        logger.warning(f"Failed to extract text from {filepath}: {e}")
        return ""


def _extract_pdf(filepath):
    from PyPDF2 import PdfReader
    from PyPDF2.errors import PdfReadError

    try:
        reader = PdfReader(filepath)
    except PdfReadError as e:
        logger.warning(f"PDF read error for {filepath}: {e}")
        return ""
    except Exception as e:
        logger.warning(f"Cannot open PDF {filepath}: {e}")
        return ""

    # Handle encrypted PDFs
    if reader.is_encrypted:
        try:
            # Try empty password (common for owner-password-only PDFs)
            if not reader.decrypt(""):
                logger.warning(f"PDF is encrypted and cannot be decrypted: {filepath}")
                return "[ENCRYPTED_PDF]"
        except Exception:
            logger.warning(f"PDF is password-protected: {filepath}")
            return "[ENCRYPTED_PDF]"

    total_pages = len(reader.pages)
    parts = []
    empty_pages = 0

    for page in reader.pages:
        try:
            text = page.extract_text()
            if text and text.strip():
                parts.append(text)
            else:
                empty_pages += 1
        except Exception as e:
            logger.warning(f"Failed to extract page from {filepath}: {e}")
            empty_pages += 1

    result = "\n".join(parts)

    # If most pages had no text, likely a scanned/image PDF
    if total_pages > 0 and empty_pages == total_pages:
        logger.warning(f"PDF appears to be scanned/image-based (0/{total_pages} pages had text): {filepath}")
        return "[SCANNED_PDF]"

    if total_pages > 0 and empty_pages > total_pages * 0.5 and parts:
        logger.info(f"PDF partially extracted ({len(parts)}/{total_pages} pages): {filepath}")

    return result


def _extract_docx(filepath):
    from docx import Document
    doc = Document(filepath)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_xlsx(filepath):
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                parts.append(" | ".join(cells))
    wb.close()
    return "\n".join(parts)


def _extract_csv(filepath):
    parts = []
    with open(filepath, "r", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            parts.append(" | ".join(row))
    return "\n".join(parts)


def _extract_txt(filepath):
    with open(filepath, "r", errors="replace") as f:
        return f.read()


# ---------------------------------------------------------------------------
# AI-powered assessment (Claude API)
# ---------------------------------------------------------------------------

def _get_anthropic_client():
    """Get Anthropic client if API key is configured."""
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return None
    try:
        import anthropic
        return anthropic.Anthropic(api_key=api_key, timeout=90.0)
    except Exception as e:
        logger.warning(f"Failed to create Anthropic client: {e}")
        return None


def _truncate_text(text, max_chars=40000):
    """Truncate text to fit within API limits while keeping meaningful content."""
    if len(text) <= max_chars:
        return text
    # Keep beginning and end (most reports have summary at start, data at end)
    half = max_chars // 2
    return text[:half] + "\n\n[... content truncated for length ...]\n\n" + text[-half:]


# Sustainability keywords for smart extraction (English + Japanese)
_SUSTAINABILITY_KEYWORDS = [
    # Core SSBJ/ISSB terms
    "sustainability", "esg", "emissions", "ghg", "greenhouse",
    "scope 1", "scope 2", "scope 3", "carbon", "climate",
    "governance", "risk management", "tcfd", "issb", "ssbj",
    "disclosure", "assurance", "environmental", "social",
    "value chain", "transition", "physical risk", "scenario",
    "intensity", "remuneration", "target", "metrics",
    "net zero", "decarboni", "renewable", "energy efficiency",
    "board", "committee", "oversight", "strategy",
    "emission factor", "activity data", "boundary", "inventory",
    "internal control", "audit trail", "reconciliation",
    "materiality", "stakeholder", "resilience",
    # Japanese terms
    "サステナビリティ", "気候", "温室効果ガス", "排出量",
    "スコープ", "ガバナンス", "リスク管理", "開示",
    "保証", "バリューチェーン", "カーボン", "脱炭素",
    "移行", "シナリオ", "取締役会", "戦略",
]


def _smart_extract(text, max_chars=60000):
    """Extract sustainability-relevant sections from large documents.

    Instead of blind first/last truncation, scores paragraphs by keyword
    relevance and keeps the most useful content. Always preserves the
    beginning of the document (executive summary / TOC area).
    """
    if len(text) <= max_chars:
        return text

    # Split into paragraphs (preserve structure)
    paragraphs = text.split("\n")

    # Always keep first ~8K chars (executive summary, TOC, intro)
    header_chars = min(8000, max_chars // 6)
    header_parts = []
    header_len = 0
    header_end_idx = 0
    for i, para in enumerate(paragraphs):
        if header_len + len(para) > header_chars:
            break
        header_parts.append(para)
        header_len += len(para) + 1
        header_end_idx = i + 1

    # Score remaining paragraphs by sustainability keyword relevance
    scored = []
    for i, para in enumerate(paragraphs[header_end_idx:], start=header_end_idx):
        para_lower = para.lower()
        if len(para.strip()) < 5:
            continue  # Skip blank/tiny lines
        score = sum(1 for kw in _SUSTAINABILITY_KEYWORDS if kw in para_lower)
        # Boost paragraphs with numbers (likely data/metrics)
        if any(c.isdigit() for c in para):
            score += 1
        scored.append((score, i, para))

    # Sort by relevance (highest first), then by original position for ties
    scored.sort(key=lambda x: (-x[0], x[1]))

    # Take highest-scoring paragraphs up to budget
    remaining_budget = max_chars - header_len
    selected = []
    for score, idx, para in scored:
        if score == 0:
            break  # No more relevant paragraphs
        if remaining_budget - len(para) < 0:
            continue  # Skip paragraphs that are too long, try smaller ones
        selected.append((idx, para))
        remaining_budget -= len(para) + 1

    # If we still have budget, add some zero-score paragraphs (context)
    for score, idx, para in scored:
        if remaining_budget <= 0:
            break
        if score == 0 and len(para.strip()) > 20:
            selected.append((idx, para))
            remaining_budget -= len(para) + 1

    # Re-sort selected paragraphs by original document order
    selected.sort(key=lambda x: x[0])

    result = "\n".join(header_parts)
    if selected:
        result += "\n\n[... relevant sections extracted ...]\n\n"
        result += "\n".join(para for _, para in selected)

    return result


def _summarize_document(text, api_key):
    """Pass 1 of two-pass assessment: extract structured sustainability content.

    Sends document text to Claude and gets back a structured summary of all
    sustainability-relevant information organized by SSBJ pillar.
    This produces a compact input for the scoring pass.
    """
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout
    import anthropic

    # Use smart extraction to get the most relevant 80K chars
    extracted = _smart_extract(text, max_chars=80000)

    system_prompt = """You are an expert sustainability disclosure analyst. Extract ALL sustainability-related information from this document into a structured summary.

Organize by these 4 categories:
1. GOVERNANCE: Board/committee oversight, management roles, policies, competencies, controls
2. STRATEGY: Climate risks/opportunities, value chain, scenario analysis, transition plans, business model impacts
3. RISK MANAGEMENT: Risk identification processes, assessment, monitoring, integration with enterprise risk
4. METRICS & TARGETS: GHG emissions (Scope 1/2/3), targets, intensity metrics, methodologies, data quality, remuneration links

For each section, extract:
- Specific facts, numbers, named entities, dates
- Named processes, policies, frameworks mentioned
- Specific methodologies or standards referenced
- Any gaps or areas where information is vague/missing

Be thorough — include ALL relevant details. This summary will be used for SSBJ compliance scoring.
If a section has no relevant content in the document, write "No relevant content found."
Keep the summary factual and concise (no commentary)."""

    user_prompt = f"""Extract all sustainability-relevant information from this document:

{extracted}

Provide a structured summary organized by the 4 categories above."""

    def _call_api():
        client = anthropic.Anthropic(api_key=api_key, timeout=40.0)
        return client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=4000,
            system=system_prompt,
            messages=[{"role": "user", "content": user_prompt}],
        )

    with ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(_call_api)
        response = future.result(timeout=45)

    return response.content[0].text.strip()


def ai_assess_all(combined_text):
    """
    Use Claude to assess all SSBJ criteria against the document text.

    For large documents (>30K chars), uses a two-pass approach:
      Pass 1: Summarize document into structured sustainability content
      Pass 2: Score 26 criteria against the summary

    For smaller documents, scores directly in a single pass.
    Runs all API calls in separate threads to prevent Gunicorn worker kills.
    Returns dict: {criterion_id: (score, evidence, notes)}
    """
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return None

    try:
        import anthropic
    except ImportError:
        return None

    # Decide strategy based on document size
    is_large = len(combined_text) > 30000

    if is_large:
        # Two-pass: summarize first, then score
        try:
            logger.info(f"Large document ({len(combined_text)} chars), using two-pass assessment")
            summary = _summarize_document(combined_text, api_key)
            scoring_text = summary
        except Exception as e:
            # If summarization fails, fall back to smart extraction
            logger.warning(f"Summarization failed ({e}), using smart extraction fallback")
            scoring_text = _smart_extract(combined_text, max_chars=60000)
    else:
        scoring_text = combined_text

    # Build criteria descriptions for the prompt
    criteria_list = []
    for c in SSBJ_CRITERIA:
        criteria_list.append(
            f"- {c['id']} ({c['pillar']} / {c['category']}): {c['requirement']}\n"
            f"  Obligation: {c['obligation']} | LA Scope: {c['la_scope']}\n"
            f"  Guidance: {c['guidance']}"
        )
    criteria_text = "\n\n".join(criteria_list)

    system_prompt = """You are an expert SSBJ/ISSB sustainability auditor. You know:
- SSBJ No.1 (IFRS S1) and No.2 (IFRS S2) requirements deeply
- Mandatory (SHALL) vs Recommended (SHOULD) vs Interpretive distinctions
- Limited assurance (ISSA 5000, replacing ISAE 3000/3410) for Scope 1 & 2 GHG, Governance, and Risk Management
- 13 essential internal controls: boundary definition, emission inventory, calculation methodology, activity data controls, emission factor management, maker-checker, audit trail, reconciliation, error tracking, management sign-off, segregation of duties, access controls, documentation
- GHG Protocol: Scope 1 (direct), Scope 2 (location + market-based), Scope 3 (15 categories)
- Japanese mandatory disclosure timeline: Phase 1 (FY ending March 2027, ≥¥3T), Phase 2 (FY ending March 2028, ≥¥1T), Phase 3 (FY ending March 2029, ≥¥500B)
- Mandatory assurance starts ONE YEAR after mandatory disclosure: Phase 1 assurance from FY ending March 2028, Phase 2 from March 2029, Phase 3 from March 2030

SCORING: 0=No evidence, 1=Mentioned only, 2=Partial processes, 3=Formal documented processes (minimum for assurance), 4=Monitored with review cycles, 5=Leading practice.
Be strict: score 3+ needs formal processes, specific methodologies, named responsibilities, concrete data. Vague mentions = 1-2.
IMPORTANT: Keep evidence and notes very brief (1 short sentence each) to stay within token limits."""

    doc_label = "DOCUMENT SUMMARY" if is_large else "DOCUMENTS"
    user_prompt = f"""Assess this content against each SSBJ criterion. Return ONLY a JSON array.

CRITERIA:
{criteria_text}

{doc_label}:
{scoring_text}

For each of the 26 criteria return: {{"id": "GOV-01", "score": 0-5, "evidence": "brief quote (1 short sentence)", "notes": "improvement needed (1 short sentence)"}}
Return ONLY valid JSON array, no other text. Keep evidence and notes concise."""

    def _call_api():
        """Run API call in thread so Gunicorn SIGABRT can't kill the worker."""
        client = anthropic.Anthropic(api_key=api_key, timeout=45.0)
        return client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=8192,
            system=system_prompt,
            messages=[{"role": "user", "content": user_prompt}],
        )

    try:
        # Run in separate thread — keep under Render's ~60s proxy timeout
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(_call_api)
            response = future.result(timeout=50)

        # Check if output was truncated (stop_reason != "end_turn")
        stop_reason = response.stop_reason
        if stop_reason != "end_turn":
            logger.warning(f"AI response truncated (stop_reason={stop_reason}), output may be incomplete")

        # Parse the response
        response_text = response.content[0].text.strip()

        # Extract JSON from response (handle potential markdown code blocks)
        if response_text.startswith("```"):
            response_text = re.sub(r'^```(?:json)?\s*\n?', '', response_text)
            response_text = re.sub(r'\n?```\s*$', '', response_text)

        # If JSON was truncated, try to salvage partial results
        try:
            results_list = json.loads(response_text)
        except json.JSONDecodeError:
            # Try to fix truncated JSON array by closing it
            fixed = response_text.rstrip().rstrip(",")
            # Find last complete object (ends with })
            last_brace = fixed.rfind("}")
            if last_brace > 0:
                fixed = fixed[:last_brace + 1] + "]"
                logger.warning("Attempting to salvage truncated JSON response")
                results_list = json.loads(fixed)
            else:
                raise

        results = {}
        for item in results_list:
            cid = item.get("id", "")
            score = int(item.get("score", 0))
            score = max(0, min(5, score))  # Clamp to 0-5
            evidence = f"[AI Assessment] {item.get('evidence', '')}"
            notes = item.get("notes", "")
            results[cid] = (score, evidence, notes)

        return results

    except FuturesTimeout:
        logger.warning("AI assessment timed out (50s)")
        raise RuntimeError("AI assessment timed out. Please try again.") from None
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse AI response as JSON: {e}")
        raise RuntimeError(f"AI returned invalid response. Please try again.") from e
    except BaseException as e:
        # Catch BaseException to handle SystemExit from Gunicorn SIGABRT
        logger.error(f"AI assessment failed: {type(e).__name__}: {e}")
        raise RuntimeError(str(e)) from e


# ---------------------------------------------------------------------------
# Parallel batch scoring (5 batches of 5 criteria)
# ---------------------------------------------------------------------------

_SCORING_SYSTEM_PROMPT = """You are an expert SSBJ/ISSB sustainability auditor. You know:
- SSBJ No.1 (IFRS S1) and No.2 (IFRS S2) requirements deeply
- Mandatory (SHALL) vs Recommended (SHOULD) vs Interpretive distinctions
- Limited assurance (ISSA 5000, replacing ISAE 3000/3410) for Scope 1 & 2 GHG, Governance, and Risk Management
- 13 essential internal controls: boundary definition, emission inventory, calculation methodology, activity data controls, emission factor management, maker-checker, audit trail, reconciliation, error tracking, management sign-off, segregation of duties, access controls, documentation
- GHG Protocol: Scope 1 (direct), Scope 2 (location + market-based), Scope 3 (15 categories)
- Japanese mandatory disclosure timeline: Phase 1 (FY ending March 2027, >=3T), Phase 2 (FY ending March 2028, >=1T), Phase 3 (FY ending March 2029, >=500B)
- Mandatory assurance starts ONE YEAR after mandatory disclosure: Phase 1 assurance from FY ending March 2028, Phase 2 from March 2029, Phase 3 from March 2030

SCORING: 0=No evidence, 1=Mentioned only, 2=Partial processes, 3=Formal documented processes (minimum for assurance), 4=Monitored with review cycles, 5=Leading practice.
Be strict: score 3+ needs formal processes, specific methodologies, named responsibilities, concrete data. Vague mentions = 1-2.
IMPORTANT: Keep evidence and notes very brief (1 short sentence each) to stay within token limits."""


def _make_batch_criteria(criteria_batch):
    """Build criteria description text for a batch prompt."""
    parts = []
    for c in criteria_batch:
        parts.append(
            f"- {c['id']} ({c['pillar']} / {c['category']}): {c['requirement']}\n"
            f"  Obligation: {c['obligation']} | LA Scope: {c['la_scope']}\n"
            f"  Guidance: {c['guidance']}"
        )
    return "\n\n".join(parts)


def _ai_assess_batch(scoring_text, criteria_batch, api_key, is_large):
    """Score a single batch of criteria against the document text.

    Runs synchronously (called inside ThreadPoolExecutor).
    Returns dict: {criterion_id: (score, evidence, notes)}
    Raises on failure so the caller can handle per-batch errors.
    """
    import anthropic

    criteria_text = _make_batch_criteria(criteria_batch)
    doc_label = "DOCUMENT SUMMARY" if is_large else "DOCUMENTS"
    ids_list = ", ".join(c["id"] for c in criteria_batch)

    user_prompt = (
        f"Assess this content against each criterion below. Return ONLY a JSON array.\n\n"
        f"CRITERIA:\n{criteria_text}\n\n"
        f"{doc_label}:\n{scoring_text}\n\n"
        f'For each of the {len(criteria_batch)} criteria ({ids_list}) return: '
        f'{{"id": "XXX-NN", "score": 0-5, "evidence": "brief (1 sentence)", '
        f'"notes": "improvement needed (1 sentence)"}}\n'
        f"Return ONLY valid JSON array, no other text."
    )

    client = anthropic.Anthropic(api_key=api_key, timeout=30.0)
    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=2000,
        system=_SCORING_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_prompt}],
    )

    text = response.content[0].text.strip()
    if text.startswith("```"):
        text = re.sub(r'^```(?:json)?\s*\n?', '', text)
        text = re.sub(r'\n?```\s*$', '', text)

    try:
        items = json.loads(text)
    except json.JSONDecodeError:
        # Try to salvage truncated JSON
        fixed = text.rstrip().rstrip(",")
        last_brace = fixed.rfind("}")
        if last_brace > 0:
            items = json.loads(fixed[:last_brace + 1] + "]")
        else:
            raise

    results = {}
    for item in items:
        cid = item.get("id", "")
        score = max(0, min(5, int(item.get("score", 0))))
        evidence = f"[AI Assessment] {item.get('evidence', '')}"
        notes = item.get("notes", "")
        results[cid] = (score, evidence, notes)
    return results


def _split_batches(criteria_list, size=5):
    """Split criteria into batches of `size`."""
    return [criteria_list[i:i + size] for i in range(0, len(criteria_list), size)]


def ai_assess_all_streaming(combined_text, batch_indices=None):
    """Generator that yields progress events during batched parallel assessment.

    Args:
        combined_text: Full document text to assess.
        batch_indices: Optional list of 0-indexed batch numbers to retry.
                       None = run all batches (normal flow).

    Yields dicts with "type" key:
        start, pass1_start, pass1_progress, pass1_done, pass1_fallback,
        pass2_start, pass2_progress, batch_done, batch_error, done

    All long-running phases yield keepalive events every few seconds to
    prevent Render's proxy from killing the SSE connection (~60s idle
    timeout).
    """
    from concurrent.futures import ThreadPoolExecutor, wait, FIRST_COMPLETED

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        results = keyword_assess_all(combined_text)
        yield {"type": "done", "results": results, "method": "keyword"}
        return

    try:
        import anthropic  # noqa: F401
    except ImportError:
        results = keyword_assess_all(combined_text)
        yield {"type": "done", "results": results, "method": "keyword"}
        return

    total = len(SSBJ_CRITERIA)

    # Check cache (only for full assessment, not partial retries)
    if batch_indices is None:
        cached = get_cached_results(combined_text)
        if cached:
            yield {"type": "cached", "scored": len(cached), "total": total}
            yield {"type": "done", "results": cached, "method": "ai_cached"}
            return

    yield {"type": "start", "total": total}

    # ---- Pass 1: summarize large documents ----
    is_large = len(combined_text) > 30000
    if is_large:
        yield {"type": "pass1_start"}
        try:
            # Run Pass 1 in a background thread so we can yield keepalive
            # events.  Without this, the SSE connection sits idle for ~40 s
            # and Render's proxy kills it.
            from concurrent.futures import ThreadPoolExecutor as _TPE

            with _TPE(max_workers=1) as pass1_exec:
                pass1_future = pass1_exec.submit(
                    _summarize_document, combined_text, api_key
                )
                elapsed = 0
                while not pass1_future.done():
                    time.sleep(2)
                    elapsed += 2
                    yield {"type": "pass1_progress", "elapsed": elapsed}
                scoring_text = pass1_future.result()  # re-raises stored exc

            yield {"type": "pass1_done"}
        except Exception as e:
            logger.warning(f"Pass 1 failed ({e}), using smart extraction")
            scoring_text = _smart_extract(combined_text, max_chars=60000)
            yield {"type": "pass1_fallback"}
    else:
        scoring_text = combined_text

    # ---- Pass 2: parallel batch scoring ----
    all_batches = _split_batches(SSBJ_CRITERIA, size=5)

    if batch_indices is not None:
        run_items = [(i, all_batches[i]) for i in batch_indices if i < len(all_batches)]
    else:
        run_items = list(enumerate(all_batches))

    yield {"type": "pass2_start", "batches": len(run_items)}

    all_results = {}
    batch_errors = {}

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {}
        for idx, batch in run_items:
            future = executor.submit(
                _ai_assess_batch, scoring_text, batch, api_key, is_large
            )
            futures[future] = idx

        # Use wait() with timeout instead of as_completed so we can
        # yield keepalive events and prevent Render proxy idle-kill.
        pending = set(futures.keys())
        while pending:
            done, pending = wait(pending, timeout=5, return_when=FIRST_COMPLETED)
            if not done:
                # No batch finished in 5s — send keepalive
                yield {
                    "type": "pass2_progress",
                    "scored": len(all_results),
                    "total": total,
                }
                continue
            for future in done:
                batch_idx = futures[future]
                batch_ids = [c["id"] for c in all_batches[batch_idx]]
                try:
                    batch_results = future.result(timeout=0)
                    all_results.update(batch_results)
                    yield {
                        "type": "batch_done",
                        "batch": batch_idx,
                        "scored": len(all_results),
                        "total": total,
                        "criteria": batch_ids,
                    }
                except Exception as e:
                    batch_errors[batch_idx] = str(e)
                    yield {
                        "type": "batch_error",
                        "batch": batch_idx,
                        "error": str(e),
                        "criteria": batch_ids,
                        "scored": len(all_results),
                        "total": total,
                    }

    # Cache only when ALL batches succeed on a full run
    if batch_indices is None and not batch_errors and all_results:
        _cache_results(combined_text, all_results)

    yield {
        "type": "done",
        "results": all_results,
        "method": "ai",
        "errors": batch_errors if batch_errors else None,
    }


# ---------------------------------------------------------------------------
# Keyword-based fallback assessment
# ---------------------------------------------------------------------------

CRITERION_KEYWORDS = {
    "GOV-01": {
        1: ["sustainability", "ESG", "oversight", "governance"],
        2: ["board", "committee", "responsible", "sustainability committee"],
        3: ["terms of reference", "mandate", "charter", "board oversight", "sustainability oversight", "reporting line"],
        4: ["regular review", "quarterly", "annual review", "monitoring", "KPI", "performance review"],
        5: ["continuous improvement", "leading practice", "integrated governance", "assurance-ready"],
    },
    "GOV-02": {
        1: ["policy", "governance", "sustainability"],
        2: ["board mandate", "corporate governance", "terms of reference"],
        3: ["sustainability policy", "governance policy", "board charter", "committee charter", "explicit reference"],
        4: ["policy review", "annual update", "policy monitoring", "compliance check"],
        5: ["integrated policy framework", "best practice", "continuous improvement"],
    },
    "GOV-03": {
        1: ["skills", "competence", "training"],
        2: ["sustainability knowledge", "board training", "expertise"],
        3: ["skills matrix", "training program", "competency framework", "sustainability expertise"],
        4: ["regular training", "competency assessment", "external advisor", "specialist"],
        5: ["continuous learning", "certified", "leading expertise"],
    },
    "GOV-04": {
        1: ["management", "role", "sustainability"],
        2: ["management responsibility", "reporting", "sustainability manager"],
        3: ["management role", "organizational chart", "reporting structure", "data owner", "role description"],
        4: ["management KPI", "performance monitoring", "regular reporting", "management review"],
        5: ["integrated management", "cross-functional", "embedded sustainability"],
    },
    "GOV-05": {
        1: ["climate", "risk", "decision"],
        2: ["climate risk", "climate factor", "investment decision"],
        3: ["climate consideration", "capital allocation", "strategic decision", "meeting minutes", "climate governance"],
        4: ["climate integration", "systematic consideration", "decision framework", "regular assessment"],
        5: ["fully integrated", "climate-aligned", "leading practice"],
    },
    "STR-01": {
        1: ["risk", "opportunity", "sustainability"],
        2: ["sustainability risk", "risk assessment", "opportunity identification"],
        3: ["risk register", "time horizon", "short-term", "medium-term", "long-term", "financial impact", "operational impact"],
        4: ["regular assessment", "risk monitoring", "quantified impact", "scenario"],
        5: ["comprehensive risk framework", "integrated risk", "dynamic assessment"],
    },
    "STR-02": {
        1: ["business model", "value chain"],
        2: ["impact assessment", "value chain risk", "business model impact"],
        3: ["value chain mapping", "dependency analysis", "business model assessment", "supply chain"],
        4: ["regular review", "dynamic assessment", "quantified dependency"],
        5: ["integrated value chain", "resilient business model"],
    },
    "STR-03": {
        1: ["financial", "impact", "sustainability"],
        2: ["financial impact", "cost", "revenue", "cash flow"],
        3: ["financial position", "cash flow", "balance sheet", "income statement", "quantified financial impact"],
        4: ["financial scenario", "stress test", "sensitivity analysis", "financial planning", "quantified impact"],
        5: ["fully quantified", "integrated financial planning", "forward-looking financial"],
    },
    "STR-04": {
        1: ["climate", "scenario"],
        2: ["scenario analysis", "climate scenario", "1.5", "2 degree"],
        3: ["climate scenario analysis", "resilience assessment", "RCP", "SSP", "IEA", "NGFS", "transition scenario", "physical scenario", "assumptions"],
        4: ["multiple scenarios", "quantified impact", "strategic implication", "time horizon", "scenario assumptions", "strategy informed", "resilience"],
        5: ["comprehensive scenario", "integrated planning", "dynamic scenario", "strategy feedback"],
    },
    "STR-05": {
        1: ["transition", "plan", "climate"],
        2: ["transition plan", "decarbonization", "net zero"],
        3: ["transition plan", "milestone", "target", "capital expenditure", "timeline", "roadmap"],
        4: ["progress tracking", "annual review", "board approved", "investment plan"],
        5: ["science-based", "SBTi", "verified", "comprehensive transition"],
    },
    "STR-06": {
        1: ["resilience", "strategy"],
        2: ["strategy resilience", "business resilience"],
        3: ["resilience assessment", "adaptability", "vulnerability", "stress test"],
        4: ["dynamic resilience", "regular reassessment", "adaptation plan"],
        5: ["fully resilient", "adaptive strategy"],
    },
    "RSK-01": {
        1: ["risk", "identification", "sustainability"],
        2: ["risk process", "risk identification", "environmental scan"],
        3: ["formal risk identification", "materiality assessment", "documented process", "risk methodology"],
        4: ["regular execution", "annual review", "stakeholder engagement", "comprehensive scan"],
        5: ["dynamic risk identification", "emerging risk", "leading practice"],
    },
    "RSK-02": {
        1: ["risk", "assessment", "monitor"],
        2: ["risk assessment", "risk prioritization", "risk monitoring"],
        3: ["risk register", "likelihood", "impact", "risk criteria", "escalation", "risk matrix"],
        4: ["regular monitoring", "risk dashboard", "KRI", "key risk indicator"],
        5: ["predictive risk", "advanced analytics", "continuous monitoring"],
    },
    "RSK-03": {
        1: ["risk management", "integration"],
        2: ["enterprise risk", "ERM", "integrated risk"],
        3: ["ERM framework", "sustainability integration", "risk appetite", "risk tolerance"],
        4: ["integrated reporting", "cross-functional", "unified framework"],
        5: ["fully integrated ERM", "leading practice"],
    },
    "RSK-04": {
        1: ["climate risk", "physical risk", "transition risk"],
        2: ["climate risk assessment", "physical risk assessment", "transition risk assessment"],
        3: ["physical risk", "transition risk", "acute", "chronic", "policy risk", "technology risk", "market risk", "reputation risk", "TCFD"],
        4: ["quantified climate risk", "regular assessment", "mitigation plan"],
        5: ["comprehensive climate risk", "integrated climate risk management"],
    },
    "RSK-05": {
        1: ["control", "data", "internal"],
        2: ["internal control", "data collection", "data quality"],
        3: ["data owner", "data collection procedure", "maker-checker", "audit trail", "reconciliation", "access control", "segregation of duties"],
        4: ["control testing", "regular review", "control monitoring", "automated control", "data validation"],
        5: ["integrated control framework", "continuous monitoring", "assurance-ready"],
    },
    "MET-01": {
        1: ["emission", "GHG", "scope 1", "greenhouse"],
        2: ["scope 1 emission", "direct emission", "GHG calculation"],
        3: ["scope 1", "direct emission", "fuel combustion", "process emission", "fugitive emission", "mobile source",
            "GHG protocol", "emission factor", "calculation methodology", "activity data", "tCO2", "CO2e"],
        4: ["verified", "third-party", "complete inventory", "regular review", "reconciliation", "sign-off"],
        5: ["assured", "continuous monitoring", "real-time", "leading methodology"],
    },
    "MET-02": {
        1: ["emission", "scope 2", "electricity", "energy"],
        2: ["scope 2 emission", "indirect emission", "purchased electricity"],
        3: ["scope 2", "location-based", "market-based", "grid emission factor", "electricity consumption",
            "utility", "energy consumption", "kWh", "MWh", "tCO2", "CO2e"],
        4: ["both approaches", "verified data", "reconciliation", "utility invoice", "regular review"],
        5: ["assured", "renewable energy certificate", "comprehensive scope 2"],
    },
    "MET-03": {
        1: ["scope 3", "value chain", "indirect"],
        2: ["scope 3 category", "upstream", "downstream", "value chain emission"],
        3: ["scope 3 category", "material category", "estimation methodology", "purchased goods", "transportation",
            "business travel", "employee commuting", "data source", "assumption"],
        4: ["comprehensive scope 3", "supplier engagement", "regular update", "data quality improvement"],
        5: ["verified scope 3", "science-based", "full value chain"],
    },
    "MET-04": {
        1: ["target", "reduction", "climate"],
        2: ["GHG target", "reduction target", "base year"],
        3: ["GHG reduction target", "base year", "target year", "milestone", "interim target", "absolute target", "intensity target"],
        4: ["progress tracking", "annual reporting", "on track", "SBTi"],
        5: ["science-based target", "net zero", "verified target"],
    },
    "MET-05": {
        1: ["industry", "metric", "sector"],
        2: ["industry metric", "sector metric", "SASB"],
        3: ["industry-specific", "SASB standard", "sector disclosure", "material metric"],
        4: ["comprehensive sector", "benchmarking", "peer comparison"],
        5: ["leading sector disclosure", "comprehensive SASB"],
    },
    "MET-06": {
        1: ["cross-industry", "carbon price", "climate metric"],
        2: ["transition risk amount", "physical risk amount", "carbon price"],
        3: ["internal carbon price", "capital deployment", "climate opportunity", "climate-related metric"],
        4: ["quantified cross-industry", "systematic measurement"],
        5: ["comprehensive cross-industry disclosure"],
    },
    "MET-07": {
        1: ["data", "quality", "accuracy"],
        2: ["data quality", "data accuracy", "data completeness"],
        3: ["data governance", "validation rule", "reconciliation", "error tracking", "data lineage", "completeness check", "data standard"],
        4: ["automated validation", "regular audit", "data quality KPI", "continuous improvement"],
        5: ["integrated data governance", "real-time validation", "leading data practice"],
    },
}


def _find_excerpts(text, keywords, max_excerpts=3):
    """Find text excerpts around keyword matches."""
    excerpts = []
    text_lower = text.lower()
    for kw in keywords:
        kw_lower = kw.lower()
        pos = text_lower.find(kw_lower)
        if pos != -1:
            start = max(0, pos - 80)
            end = min(len(text), pos + len(kw) + 80)
            snippet = text[start:end].strip()
            snippet = re.sub(r'\s+', ' ', snippet)
            if snippet and snippet not in excerpts:
                excerpts.append(f"...{snippet}...")
            if len(excerpts) >= max_excerpts:
                break
    return excerpts


def keyword_assess_all(combined_text):
    """
    Keyword-based fallback assessment for all SSBJ criteria.

    Returns dict: {criterion_id: (score, evidence, notes)}
    """
    results = {}
    text_lower = combined_text.lower()

    for criterion in SSBJ_CRITERIA:
        cid = criterion["id"]
        keywords_by_level = CRITERION_KEYWORDS.get(cid, {})

        if not keywords_by_level:
            results[cid] = (0, "", "No assessment keywords defined for this criterion.")
            continue

        best_level = 0
        all_matched = []
        all_excerpts = []

        for level in sorted(keywords_by_level.keys()):
            kws = keywords_by_level[level]
            matched = [kw for kw in kws if kw.lower() in text_lower]
            if matched:
                best_level = level
                all_matched.extend(matched)
                excerpts = _find_excerpts(combined_text, matched, max_excerpts=2)
                all_excerpts.extend(excerpts)

        if best_level == 0:
            results[cid] = (0, "", "No relevant content found in uploaded documents.")
            continue

        unique_matched = list(dict.fromkeys(all_matched))
        evidence = f"[Keyword match] Keywords found: {', '.join(unique_matched[:10])}"
        if all_excerpts:
            unique_excerpts = list(dict.fromkeys(all_excerpts))
            evidence += "\n\nRelevant excerpts:\n" + "\n".join(unique_excerpts[:4])

        notes = f"Maturity level {best_level} based on keyword matching (not AI analysis)."
        notes += f"\nGuidance: {criterion['guidance']}"

        results[cid] = (best_level, evidence, notes)

    return results


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def auto_assess_all(combined_text):
    """
    Synchronous entry point for auto-assessment.

    Uses the batched parallel pipeline (ai_assess_all_streaming) internally,
    consuming all events and returning the final results.
    Includes content-hash caching — repeat calls return instantly.

    Returns (results_dict, method_used, error_message).
    - results_dict: {criterion_id: (score, evidence, notes)}
    - method_used: "ai" | "ai_cached" | "keyword"
    - error_message: None if success, or string describing failures
    """
    results = {}
    method = "keyword"
    error_msg = None

    for event in ai_assess_all_streaming(combined_text):
        if event["type"] == "done":
            results = event.get("results", {})
            method = event.get("method", "ai")
            errors = event.get("errors")
            if errors:
                error_msg = "; ".join(f"Batch {k}: {v}" for k, v in errors.items())

    if not results:
        # Total failure — fall back to keyword
        results = keyword_assess_all(combined_text)
        method = "keyword"

    return results, method, error_msg
