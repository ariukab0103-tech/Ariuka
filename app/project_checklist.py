"""
Project Execution Checklist Generator

Generates a structured, phase-by-phase checklist for companies that received the
SSBJ gap analysis report. Bridges the gap between "here's your assessment" and
"here's your project plan" by auto-generating:

1. Phase-by-phase task checklist with concrete deliverables
2. Evidence & documentation tracker (what the auditor will request)
3. Resource & budget indicators per phase
4. Assurance preparation milestones with gate reviews
5. Year 2+ forward-look for deferred items

All data is derived from existing assessment scores, RACI assignments,
roadmap phases, and relief advisor outputs.
"""

from datetime import date
from app.ssbj_criteria import SSBJ_CRITERIA


# Evidence artifacts the auditor will request per criterion (ISSA 5000 / ISAE 3410)
_EVIDENCE_MAP = {
    "GOV-01": [
        {"document": "Board/committee charter (terms of reference) with sustainability mandate", "format": "PDF/Word"},
        {"document": "Board meeting minutes showing sustainability agenda items", "format": "PDF"},
        {"document": "Committee membership list with appointment dates", "format": "PDF/Excel"},
    ],
    "GOV-02": [
        {"document": "Management role descriptions (data owner, reviewer, approver)", "format": "PDF/Word"},
        {"document": "Organizational chart showing sustainability reporting lines", "format": "PDF"},
        {"document": "Sign-off authorization matrix for sustainability data", "format": "Excel"},
    ],
    "GOV-03": [
        {"document": "Board skills matrix with ESG/sustainability competencies", "format": "Excel"},
        {"document": "Training records for sustainability-related sessions", "format": "PDF"},
    ],
    "GOV-04": [
        {"document": "Board meeting minutes discussing sustainability in strategy", "format": "PDF"},
        {"document": "Strategic planning documents referencing climate/sustainability risks", "format": "PDF/Word"},
    ],
    "GOV-05": [
        {"document": "Board resolution on climate target-setting / approval of targets", "format": "PDF"},
        {"document": "Target oversight framework document", "format": "PDF/Word"},
    ],
    "STR-01": [
        {"document": "Time horizon definitions (short/medium/long-term)", "format": "Word/PDF"},
        {"document": "Climate/sustainability risk register", "format": "Excel"},
        {"document": "Value chain assessment boundary document", "format": "Word/PDF"},
        {"document": "Hotspot identification analysis", "format": "Excel/PDF"},
    ],
    "STR-02": [
        {"document": "Value chain map (upstream to downstream)", "format": "Diagram/PDF"},
        {"document": "Assessment boundary disclosure (included/excluded and why)", "format": "Word/PDF"},
        {"document": "Business model impact assessment", "format": "Word/PDF"},
    ],
    "STR-03": [
        {"document": "Financial effects analysis (position, performance, cash flows)", "format": "Word/PDF"},
        {"document": "Quantification methodology (or explanation of why qualitative)", "format": "Word/PDF"},
    ],
    "STR-04": [
        {"document": "Scenario selection rationale and source citations", "format": "Word/PDF"},
        {"document": "Assumptions and parameters per scenario", "format": "Excel/Word"},
        {"document": "Strategy feedback narrative (how analysis informed strategy)", "format": "Word/PDF"},
        {"document": "Resilience assessment conclusion", "format": "Word/PDF"},
    ],
    "STR-05": [
        {"document": "Transition plan document (even if directional)", "format": "Word/PDF"},
        {"document": "Board approval of transition commitments", "format": "PDF"},
    ],
    "STR-06": [
        {"document": "Resilience assessment (climate resilience covered under STR-04)", "format": "Word/PDF"},
    ],
    "STR-07": [
        {"document": "Mapping table: sustainability disclosures to financial statement line items", "format": "Excel"},
    ],
    "RSK-01": [
        {"document": "Risk identification methodology document", "format": "Word/PDF"},
        {"document": "Climate risk register (physical + transition)", "format": "Excel"},
    ],
    "RSK-02": [
        {"document": "Risk assessment criteria and scoring methodology", "format": "Word/PDF"},
        {"document": "Risk assessment results with prioritization", "format": "Excel"},
    ],
    "RSK-03": [
        {"document": "Risk response strategies per identified risk", "format": "Word/PDF"},
        {"document": "Mitigation action plan with timelines", "format": "Excel"},
    ],
    "RSK-04": [
        {"document": "ERM framework showing climate risk integration", "format": "Word/PDF"},
        {"document": "Cross-reference: sustainability risk register to enterprise risk register", "format": "Excel"},
    ],
    "RSK-05": [
        {"document": "Internal controls framework document (data governance)", "format": "Word/PDF"},
        {"document": "Data collection procedures (step-by-step)", "format": "Word/PDF"},
        {"document": "Maker-checker review log (who reviewed, when, sign-off)", "format": "Excel"},
        {"document": "Data reconciliation checklist", "format": "Excel"},
        {"document": "Audit trail (data change log)", "format": "Excel/System"},
    ],
    "MET-01": [
        {"document": "Scope 1 emission source inventory (all facilities)", "format": "Excel"},
        {"document": "Calculation methodology document (emission factors, sources)", "format": "Word/PDF"},
        {"document": "Activity data (fuel invoices, meter readings)", "format": "Original + Excel"},
        {"document": "Calculation spreadsheet with review sign-off", "format": "Excel"},
        {"document": "Emission factor sources and version documentation", "format": "PDF/Excel"},
    ],
    "MET-02": [
        {"document": "Scope 2 facility electricity consumption data", "format": "Excel + invoices"},
        {"document": "Grid emission factors with area-specific sources", "format": "Excel"},
        {"document": "Location-based calculation spreadsheet", "format": "Excel"},
        {"document": "Market-based calculation (if green energy purchased)", "format": "Excel"},
    ],
    "MET-03": [
        {"document": "Scope 3 category materiality assessment (all 15 categories)", "format": "Excel"},
        {"document": "Supplier engagement plan and correspondence", "format": "Word/PDF"},
        {"document": "Data collection plan per category", "format": "Excel"},
    ],
    "MET-04": [
        {"document": "Climate targets with base year, target year, methodology", "format": "Word/PDF"},
        {"document": "Board approval of targets", "format": "PDF"},
    ],
    "MET-05": [
        {"document": "Industry-specific metrics disclosure (if applicable)", "format": "Word/PDF"},
    ],
    "MET-06": [
        {"document": "Climate-related financial amounts or size information", "format": "Word/PDF"},
    ],
    "MET-07": [
        {"document": "Data governance policy", "format": "Word/PDF"},
        {"document": "Data flow diagram (source to disclosure)", "format": "Diagram/PDF"},
        {"document": "Validation rules and procedures", "format": "Word/PDF"},
        {"document": "Error log and correction records", "format": "Excel"},
        {"document": "Completeness check records", "format": "Excel"},
    ],
    "MET-08": [
        {"document": "GHG intensity calculation and denominator selection rationale", "format": "Excel"},
    ],
    "MET-09": [
        {"document": "Remuneration policy referencing climate/sustainability metrics", "format": "PDF"},
        {"document": "Board resolution on climate-linked remuneration", "format": "PDF"},
    ],
}

# Resource/effort indicators per criterion for gap closure
_EFFORT_MAP = {
    "GOV-01": {"effort_days": "3-5", "skills": ["Board secretary", "ESG lead"], "external": False},
    "GOV-02": {"effort_days": "3-5", "skills": ["ESG lead", "HR"], "external": False},
    "GOV-03": {"effort_days": "2-3", "skills": ["HR", "ESG lead"], "external": False},
    "GOV-04": {"effort_days": "3-5", "skills": ["Board secretary", "Strategy"], "external": False},
    "GOV-05": {"effort_days": "3-5", "skills": ["Board secretary", "ESG lead"], "external": False},
    "STR-01": {"effort_days": "10-20", "skills": ["Risk management", "ESG lead", "Operations"], "external": True},
    "STR-02": {"effort_days": "10-20", "skills": ["ESG lead", "Operations", "Procurement"], "external": True},
    "STR-03": {"effort_days": "10-15", "skills": ["Finance", "ESG lead"], "external": True},
    "STR-04": {"effort_days": "15-25", "skills": ["ESG lead", "Risk management", "External consultant"], "external": True},
    "STR-05": {"effort_days": "10-15", "skills": ["ESG lead", "Strategy", "Operations"], "external": True},
    "STR-06": {"effort_days": "5-10", "skills": ["ESG lead", "Risk management"], "external": False},
    "STR-07": {"effort_days": "5-10", "skills": ["Finance", "ESG lead", "IR"], "external": False},
    "RSK-01": {"effort_days": "5-10", "skills": ["Risk management", "ESG lead"], "external": False},
    "RSK-02": {"effort_days": "5-10", "skills": ["Risk management", "ESG lead"], "external": False},
    "RSK-03": {"effort_days": "5-10", "skills": ["Risk management", "ESG lead", "Operations"], "external": False},
    "RSK-04": {"effort_days": "5-10", "skills": ["Risk management", "ESG lead"], "external": False},
    "RSK-05": {"effort_days": "15-30", "skills": ["Finance/Accounting", "IT", "ESG lead"], "external": True},
    "MET-01": {"effort_days": "15-25", "skills": ["Operations", "ESG lead", "External verifier"], "external": True},
    "MET-02": {"effort_days": "10-15", "skills": ["Operations", "ESG lead", "Facilities"], "external": True},
    "MET-03": {"effort_days": "20-40", "skills": ["Procurement", "ESG lead", "External consultant"], "external": True},
    "MET-04": {"effort_days": "5-10", "skills": ["ESG lead", "Board secretary"], "external": False},
    "MET-05": {"effort_days": "3-5", "skills": ["ESG lead", "IR"], "external": False},
    "MET-06": {"effort_days": "5-10", "skills": ["Finance", "ESG lead"], "external": False},
    "MET-07": {"effort_days": "15-25", "skills": ["Finance/Accounting", "IT", "ESG lead"], "external": True},
    "MET-08": {"effort_days": "3-5", "skills": ["ESG lead", "Finance"], "external": False},
    "MET-09": {"effort_days": "5-10", "skills": ["HR", "Board secretary", "Legal"], "external": False},
}

# Budget categories
_BUDGET_CATEGORIES = {
    "internal_fte": "Internal staff time",
    "ghg_software": "GHG calculation software / tools",
    "external_consultant": "External ESG consultant",
    "assurance_fees": "Assurance engagement fees",
    "training": "Staff training & capacity building",
    "data_systems": "Data management systems",
    "supplier_engagement": "Supplier engagement program",
}


def _classify_phase(criterion, score, relief_applicable, la_scope, months_remaining):
    """Determine which roadmap phase a criterion's gap closure belongs to."""
    if score is None:
        return 1  # Unscored = start immediately

    if la_scope == "in_scope" and score < 3:
        # LA-critical gaps: front-load
        if months_remaining <= 12:
            return 1
        return 2

    if score == 0:
        return 1  # Not started = immediate
    elif score == 1:
        if la_scope == "in_scope":
            return 2
        return 3
    elif score == 2:
        if la_scope == "in_scope":
            return 3
        return 4
    else:
        # Score 3+ but could improve
        return 5

    return 3  # fallback


def generate_checklist(assessment, responses, roadmap_data=None, raci_data=None, relief_data=None):
    """
    Generate a project execution checklist from assessment data.

    Args:
        assessment: Assessment model instance
        responses: dict of {criterion_id: Response}
        roadmap_data: dict from generate_roadmap() (optional, generated if None)
        raci_data: dict from generate_raci() (optional, generated if None)
        relief_data: dict from generate_relief_plan() (optional, generated if None)

    Returns dict with:
        - phases: list of phase dicts with tasks, evidence, resources
        - evidence_tracker: list of all evidence items with status
        - budget_summary: estimated budget categories
        - gate_reviews: list of milestone gates
        - year2_prep: deferred items requiring Year 1 groundwork
        - summary: overall stats
    """
    criteria_map = {c["id"]: c for c in SSBJ_CRITERIA}

    # Generate dependent data if not provided
    if roadmap_data is None:
        from app.roadmap import generate_roadmap
        scored = [r for r in responses.values() if r.score is not None]
        if scored:
            roadmap_data = generate_roadmap(assessment, list(scored))

    if raci_data is None:
        from app.raci import generate_raci
        raci_data = generate_raci(assessment, responses)

    if relief_data is None:
        from app.relief_advisor import generate_relief_plan
        relief_data = generate_relief_plan(assessment, responses)

    months_remaining = roadmap_data["months_remaining"] if roadmap_data else 18

    # Build RACI lookup: criterion_id -> {dept_code: role}
    raci_lookup = {}
    if raci_data:
        for row in raci_data.get("criteria", []):
            raci_lookup[row["id"]] = row.get("dept_roles", {})

    # Relief lookup
    relief_lookup = {}
    if relief_data:
        for item in relief_data.get("relief_items", []):
            relief_lookup[item["criterion_id"]] = item

    # --- Build per-criterion task items ---
    all_tasks = []
    total_effort_min = 0
    total_effort_max = 0
    needs_external = False

    for c in SSBJ_CRITERIA:
        cid = c["id"]
        resp = responses.get(cid)
        score = resp.score if resp and resp.score is not None else None
        evidence_text = (resp.evidence or "") if resp else ""

        is_gap = score is not None and score < 3
        is_unscored = score is None
        target_score = 3 if c["la_scope"] == "in_scope" else 3

        # Phase assignment
        phase_num = _classify_phase(
            c, score,
            cid in relief_lookup and relief_lookup[cid].get("applicable"),
            c["la_scope"],
            months_remaining,
        )

        # Evidence items
        evidence_items = []
        for ev in _EVIDENCE_MAP.get(cid, []):
            # Check if evidence text mentions key words from the document name
            doc_keywords = ev["document"].lower().split()[:3]
            mentioned = any(kw in evidence_text.lower() for kw in doc_keywords if len(kw) > 3)
            evidence_items.append({
                "document": ev["document"],
                "format": ev["format"],
                "status": "likely_exists" if mentioned and score and score >= 3 else (
                    "in_progress" if mentioned else "not_started"
                ),
            })

        # RACI for this criterion
        raci_roles = raci_lookup.get(cid, {})
        responsible = [code for code, role in raci_roles.items() if role == "R"]
        accountable = [code for code, role in raci_roles.items() if role == "A"]

        # Effort
        effort = _EFFORT_MAP.get(cid, {"effort_days": "5-10", "skills": [], "external": False})
        if is_gap or is_unscored:
            parts = effort["effort_days"].split("-")
            total_effort_min += int(parts[0])
            total_effort_max += int(parts[1]) if len(parts) > 1 else int(parts[0])
            if effort["external"]:
                needs_external = True

        # Relief info
        relief_info = relief_lookup.get(cid)
        can_defer = relief_info and relief_info.get("is_deferral") and relief_info.get("applicable")
        can_simplify = relief_info and relief_info.get("applicable") and not relief_info.get("is_deferral")

        task = {
            "criterion_id": cid,
            "pillar": c["pillar"],
            "category": c["category"],
            "standard": c["standard"],
            "obligation": c["obligation"],
            "la_scope": c["la_scope"],
            "la_priority": c.get("la_priority", ""),
            "score": score,
            "target_score": target_score,
            "is_gap": is_gap,
            "is_unscored": is_unscored,
            "phase": phase_num,
            "deliverable": c.get("minimum_action", ""),
            "evidence_items": evidence_items,
            "responsible": responsible,
            "accountable": accountable,
            "effort_days": effort["effort_days"],
            "skills_needed": effort["skills"],
            "external_help": effort["external"],
            "can_defer": can_defer,
            "can_simplify": can_simplify,
            "relief_note": relief_info["relief"] if relief_info and relief_info.get("applicable") else None,
            "year2_req": relief_info["year2_requirement"] if relief_info else None,
        }
        all_tasks.append(task)

    # --- Group tasks into phases ---
    phase_defs = [
        {"number": 1, "title": "Immediate Actions", "subtitle": "Weeks 1-4: Foundation & critical gaps",
         "icon": "bi-lightning", "color": "danger"},
        {"number": 2, "title": "Governance & Controls Setup", "subtitle": "Months 1-3: Policies, roles, frameworks",
         "icon": "bi-building", "color": "primary"},
        {"number": 3, "title": "Data & Process Build", "subtitle": "Months 3-6: Systems, calculations, evidence",
         "icon": "bi-gear", "color": "warning"},
        {"number": 4, "title": "Dry Run & Review", "subtitle": "Months 6-9: Internal review, mock audit prep",
         "icon": "bi-clipboard-check", "color": "info"},
        {"number": 5, "title": "Polish & Assurance Prep", "subtitle": "Months 9-12: Final prep, auditor engagement",
         "icon": "bi-shield-check", "color": "success"},
    ]

    phases = []
    for pdef in phase_defs:
        phase_tasks = [t for t in all_tasks if t["phase"] == pdef["number"]]
        gap_tasks = [t for t in phase_tasks if t["is_gap"] or t["is_unscored"]]
        ok_tasks = [t for t in phase_tasks if not t["is_gap"] and not t["is_unscored"]]

        phase_effort_min = sum(int(t["effort_days"].split("-")[0]) for t in gap_tasks)
        parts_max = [t["effort_days"].split("-") for t in gap_tasks]
        phase_effort_max = sum(int(p[1]) if len(p) > 1 else int(p[0]) for p in parts_max)

        phases.append({
            **pdef,
            "tasks": phase_tasks,
            "gap_tasks": gap_tasks,
            "ok_tasks": ok_tasks,
            "task_count": len(phase_tasks),
            "gap_count": len(gap_tasks),
            "effort_range": f"{phase_effort_min}-{phase_effort_max}" if gap_tasks else "0",
        })

    # --- Evidence tracker (flat list across all criteria) ---
    evidence_tracker = []
    for task in all_tasks:
        for ev in task["evidence_items"]:
            evidence_tracker.append({
                "criterion_id": task["criterion_id"],
                "pillar": task["pillar"],
                "category": task["category"],
                "la_scope": task["la_scope"],
                "document": ev["document"],
                "format": ev["format"],
                "status": ev["status"],
                "phase": task["phase"],
                "responsible": task["responsible"],
            })

    # --- Budget summary ---
    gap_count = sum(1 for t in all_tasks if t["is_gap"])
    la_gap_count = sum(1 for t in all_tasks if t["is_gap"] and t["la_scope"] == "in_scope")
    external_items = sum(1 for t in all_tasks if t["is_gap"] and t["external_help"])

    budget_items = []
    budget_items.append({
        "category": "Internal staff time",
        "estimate": f"{total_effort_min}-{total_effort_max} person-days",
        "note": f"Across {gap_count} gap items. Assumes dedicated project team.",
    })
    if external_items > 0:
        budget_items.append({
            "category": "External ESG consultant",
            "estimate": "¥5M-¥15M",
            "note": f"{external_items} items may need external expertise (scenario analysis, GHG calculation, value chain mapping).",
        })
    budget_items.append({
        "category": "Assurance engagement fees",
        "estimate": "¥5M-¥30M+",
        "note": "Limited assurance for Scope 1 & 2, Governance, Risk Management. Varies by company size and complexity.",
    })
    if any(t["is_gap"] and t["criterion_id"] in ("MET-01", "MET-02", "MET-07", "RSK-05") for t in all_tasks):
        budget_items.append({
            "category": "GHG calculation software / data systems",
            "estimate": "¥1M-¥10M/year",
            "note": "Consider: Zeroboard, Persefoni, booost, or enhanced spreadsheet approach for Year 1.",
        })
    budget_items.append({
        "category": "Training & capacity building",
        "estimate": "¥0.5M-¥2M",
        "note": "Board sustainability briefing, GHG accounting training, ISSA 5000 awareness.",
    })

    # --- Gate reviews (milestone checkpoints) ---
    gate_reviews = [
        {
            "gate": "G1: Project Kickoff",
            "timing": "Week 1",
            "criteria": "Budget approved, project team assigned, gap analysis reviewed by management",
            "owner": "Project Sponsor",
            "phase": 1,
        },
        {
            "gate": "G2: Governance Ready",
            "timing": "Month 3",
            "criteria": "Board charter amended, management roles assigned, sustainability committee established",
            "owner": "ESG Lead / Board Secretary",
            "phase": 2,
        },
        {
            "gate": "G3: Data Collection Complete",
            "timing": "Month 6",
            "criteria": "Scope 1 & 2 data collected for all sites, internal controls documented, data quality checks in place",
            "owner": "ESG Lead / Data Owner",
            "phase": 3,
        },
        {
            "gate": "G4: Dry Run Complete",
            "timing": "Month 9",
            "criteria": "Draft disclosure complete, internal review passed, evidence binders organized",
            "owner": "ESG Lead",
            "phase": 4,
        },
        {
            "gate": "G5: Assurance Ready",
            "timing": "Month 12",
            "criteria": "Assurance provider selected, pre-engagement review complete, all LA-scope items at score 3+",
            "owner": "Project Sponsor",
            "phase": 5,
        },
    ]

    # --- Year 2+ preparation items ---
    year2_prep = []
    for task in all_tasks:
        if task["can_defer"]:
            year2_prep.append({
                "criterion_id": task["criterion_id"],
                "category": task["category"],
                "relief_note": task["relief_note"],
                "year2_req": task["year2_req"],
                "year1_action": "Begin groundwork even though calculation can be deferred. Year 2 auditors will ask what you did in Year 1.",
            })

    # Items not deferred but need Year 2 expansion
    for item in relief_data.get("relief_items", []) if relief_data else []:
        if not item.get("is_deferral") and item.get("applicable") and item.get("year2_requirement"):
            if not any(y["criterion_id"] == item["criterion_id"] for y in year2_prep):
                year2_prep.append({
                    "criterion_id": item["criterion_id"],
                    "category": item["category"],
                    "relief_note": item["relief"],
                    "year2_req": item["year2_requirement"],
                    "year1_action": f"Year 1 simplified approach: {item['what_you_must_do']}",
                })

    # --- Summary stats ---
    summary = {
        "total_criteria": len(SSBJ_CRITERIA),
        "total_gaps": gap_count,
        "la_gaps": la_gap_count,
        "total_evidence_items": len(evidence_tracker),
        "evidence_not_started": sum(1 for e in evidence_tracker if e["status"] == "not_started"),
        "evidence_in_progress": sum(1 for e in evidence_tracker if e["status"] == "in_progress"),
        "evidence_likely_exists": sum(1 for e in evidence_tracker if e["status"] == "likely_exists"),
        "total_effort_range": f"{total_effort_min}-{total_effort_max}",
        "needs_external_help": needs_external,
        "external_items_count": external_items,
        "year2_deferred_count": len([y for y in year2_prep if "deferred" in (y.get("relief_note") or "").lower()]),
        "year2_expansion_count": len(year2_prep),
        "months_remaining": months_remaining,
        "gate_count": len(gate_reviews),
    }

    return {
        "phases": phases,
        "evidence_tracker": evidence_tracker,
        "budget_summary": budget_items,
        "gate_reviews": gate_reviews,
        "year2_prep": year2_prep,
        "summary": summary,
    }


# Department code -> full name mapping (matches raci.py DEPARTMENTS)
_DEPT_NAMES = {
    "board": "Board / Sustainability Committee",
    "esg": "Sustainability / ESG Office",
    "finance": "Finance / Accounting",
    "legal": "Legal / Compliance",
    "risk": "Risk Management",
    "ops": "Operations / Manufacturing",
    "hr": "HR / General Affairs",
    "ir": "IR / Communications",
    "it": "IT / Systems",
    "procurement": "Procurement / Supply Chain",
}


def generate_excel(checklist_data, assessment_title="", entity_name="", fiscal_year="", review_data=None):
    """
    Generate an Excel workbook from checklist data.

    Args:
        checklist_data: dict from generate_checklist()
        assessment_title: str
        entity_name: str
        fiscal_year: str

    Returns:
        openpyxl.Workbook ready to be saved or streamed
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Color definitions
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    phase_fills = {
        1: PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),  # red
        2: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),  # blue
        3: PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),  # yellow
        4: PatternFill(start_color="CFFAFE", end_color="CFFAFE", fill_type="solid"),  # cyan
        5: PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),  # green
    }
    la_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
    gap_font = Font(color="DC2626", bold=True)
    ok_font = Font(color="16A34A")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    def _apply_header(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = thin_border

    def _auto_width(ws, min_width=8, max_width=50):
        for col_cells in ws.columns:
            max_len = min_width
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    longest = max(len(line) for line in lines) if lines else 0
                    max_len = max(max_len, min(longest + 2, max_width))
            ws.column_dimensions[col_letter].width = max_len

    # ===== Sheet 1: Phase Checklist =====
    ws1 = wb.active
    ws1.title = "Phase Checklist"
    ws1.sheet_properties.tabColor = "3B82F6"

    # Title row
    ws1.merge_cells("A1:K1")
    title_cell = ws1["A1"]
    title_cell.value = f"SSBJ Project Execution Checklist — {entity_name} — {fiscal_year}"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    ws1.merge_cells("A2:K2")
    ws1["A2"].value = f"Assessment: {assessment_title} | Gaps: {checklist_data['summary']['total_gaps']} | LA-Critical: {checklist_data['summary']['la_gaps']} | Effort: {checklist_data['summary']['total_effort_range']} person-days"
    ws1["A2"].font = Font(size=10, italic=True, color="6B7280")
    ws1["A2"].alignment = Alignment(horizontal="center")

    headers = ["Done", "Phase", "Criterion", "Category", "Pillar", "LA Scope",
               "Current Score", "Target", "Effort (days)", "Responsible", "Accountable",
               "Relief Available", "Key Deliverable"]
    row = 4
    for ci, h in enumerate(headers, 1):
        ws1.cell(row=row, column=ci, value=h)
    _apply_header(ws1, row, len(headers))
    row += 1

    for phase in checklist_data["phases"]:
        for task in phase["gap_tasks"]:
            r = [code for code in task["responsible"]]
            a = [code for code in task["accountable"]]
            resp_names = ", ".join(_DEPT_NAMES.get(d, d) for d in r)
            acct_names = ", ".join(_DEPT_NAMES.get(d, d) for d in a)

            relief = ""
            if task["can_defer"]:
                relief = "Deferral OK"
            elif task["can_simplify"]:
                relief = "Simplified OK"

            ws1.cell(row=row, column=1, value="").border = thin_border
            ws1.cell(row=row, column=2, value=f"P{phase['number']}: {phase['title']}").border = thin_border
            ws1.cell(row=row, column=3, value=task["criterion_id"]).border = thin_border
            ws1.cell(row=row, column=4, value=task["category"]).border = thin_border
            ws1.cell(row=row, column=5, value=task["pillar"]).border = thin_border
            ws1.cell(row=row, column=6, value=task["la_scope"].replace("_", " ").title()).border = thin_border
            score_cell = ws1.cell(row=row, column=7, value=task["score"] if task["score"] is not None else "Not scored")
            score_cell.border = thin_border
            if task["score"] is not None and task["score"] < 3:
                score_cell.font = gap_font
            ws1.cell(row=row, column=8, value=task["target_score"]).border = thin_border
            ws1.cell(row=row, column=9, value=task["effort_days"]).border = thin_border
            ws1.cell(row=row, column=10, value=resp_names).border = thin_border
            ws1.cell(row=row, column=11, value=acct_names).border = thin_border
            ws1.cell(row=row, column=12, value=relief).border = thin_border
            deliv_cell = ws1.cell(row=row, column=13, value=task["deliverable"][:500] if task["deliverable"] else "")
            deliv_cell.border = thin_border
            deliv_cell.alignment = wrap

            # Phase color fill on column B
            ws1.cell(row=row, column=2).fill = phase_fills.get(phase["number"], PatternFill())

            # LA scope highlight
            if task["la_scope"] == "in_scope":
                ws1.cell(row=row, column=6).fill = la_fill
                ws1.cell(row=row, column=6).font = Font(color="DC2626", bold=True)

            for ci in range(1, len(headers) + 1):
                ws1.cell(row=row, column=ci).alignment = wrap

            row += 1

    _auto_width(ws1)
    ws1.column_dimensions["M"].width = 60  # Deliverable column wider
    ws1.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{row - 1}"
    ws1.freeze_panes = "A5"

    # ===== Sheet 2: Evidence Tracker =====
    ws2 = wb.create_sheet("Evidence Tracker")
    ws2.sheet_properties.tabColor = "EAB308"

    ev_headers = ["Done", "Criterion", "Category", "Pillar", "LA Scope",
                  "Document Required", "Format", "Status", "Phase", "Owner", "Notes"]
    for ci, h in enumerate(ev_headers, 1):
        ws2.cell(row=1, column=ci, value=h)
    _apply_header(ws2, 1, len(ev_headers))

    row = 2
    for ev in checklist_data["evidence_tracker"]:
        resp_names = ", ".join(_DEPT_NAMES.get(d, d) for d in ev["responsible"])
        status_text = ev["status"].replace("_", " ").title()

        ws2.cell(row=row, column=1, value="").border = thin_border
        ws2.cell(row=row, column=2, value=ev["criterion_id"]).border = thin_border
        ws2.cell(row=row, column=3, value=ev["category"]).border = thin_border
        ws2.cell(row=row, column=4, value=ev["pillar"]).border = thin_border
        la_cell = ws2.cell(row=row, column=5, value=ev["la_scope"].replace("_", " ").title())
        la_cell.border = thin_border
        if ev["la_scope"] == "in_scope":
            la_cell.fill = la_fill
            la_cell.font = Font(color="DC2626", bold=True)
        ws2.cell(row=row, column=6, value=ev["document"]).border = thin_border
        ws2.cell(row=row, column=7, value=ev["format"]).border = thin_border
        status_cell = ws2.cell(row=row, column=8, value=status_text)
        status_cell.border = thin_border
        if ev["status"] == "not_started":
            status_cell.font = Font(color="DC2626")
        elif ev["status"] == "likely_exists":
            status_cell.font = ok_font
        ws2.cell(row=row, column=9, value=f"P{ev['phase']}").border = thin_border
        ws2.cell(row=row, column=10, value=resp_names).border = thin_border
        ws2.cell(row=row, column=11, value="").border = thin_border  # Notes column for user

        for ci in range(1, len(ev_headers) + 1):
            ws2.cell(row=row, column=ci).alignment = wrap

        row += 1

    _auto_width(ws2)
    ws2.column_dimensions["F"].width = 55
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(ev_headers))}{row - 1}"
    ws2.freeze_panes = "A2"

    # ===== Sheet 3: Budget & Resources =====
    ws3 = wb.create_sheet("Budget & Resources")
    ws3.sheet_properties.tabColor = "8B5CF6"

    budget_headers = ["Category", "Estimate", "Notes"]
    for ci, h in enumerate(budget_headers, 1):
        ws3.cell(row=1, column=ci, value=h)
    _apply_header(ws3, 1, len(budget_headers))

    row = 2
    for item in checklist_data["budget_summary"]:
        ws3.cell(row=row, column=1, value=item["category"]).border = thin_border
        ws3.cell(row=row, column=2, value=item["estimate"]).border = thin_border
        ws3.cell(row=row, column=3, value=item["note"]).border = thin_border
        for ci in range(1, 4):
            ws3.cell(row=row, column=ci).alignment = wrap
        row += 1

    # Blank row then resource summary
    row += 1
    ws3.cell(row=row, column=1, value="Resource Summary").font = Font(bold=True, size=11)
    row += 1
    ws3.cell(row=row, column=1, value="Total Effort")
    ws3.cell(row=row, column=2, value=f"{checklist_data['summary']['total_effort_range']} person-days")
    row += 1
    ws3.cell(row=row, column=1, value="External Help Needed")
    ws3.cell(row=row, column=2, value=f"Yes ({checklist_data['summary']['external_items_count']} items)" if checklist_data['summary']['needs_external_help'] else "No")
    row += 1
    ws3.cell(row=row, column=1, value="Total Gap Items")
    ws3.cell(row=row, column=2, value=checklist_data["summary"]["total_gaps"])
    row += 1
    ws3.cell(row=row, column=1, value="LA-Critical Gaps")
    ws3.cell(row=row, column=2, value=checklist_data["summary"]["la_gaps"])
    row += 1
    ws3.cell(row=row, column=1, value="Months Remaining")
    ws3.cell(row=row, column=2, value=checklist_data["summary"]["months_remaining"])

    _auto_width(ws3)
    ws3.column_dimensions["C"].width = 60

    # ===== Sheet 4: Gate Reviews =====
    ws4 = wb.create_sheet("Gate Reviews")
    ws4.sheet_properties.tabColor = "06B6D4"

    gate_headers = ["Gate", "Timing", "Pass Criteria", "Owner", "Status", "Actual Date", "Notes"]
    for ci, h in enumerate(gate_headers, 1):
        ws4.cell(row=1, column=ci, value=h)
    _apply_header(ws4, 1, len(gate_headers))

    row = 2
    for gate in checklist_data["gate_reviews"]:
        ws4.cell(row=row, column=1, value=gate["gate"]).border = thin_border
        ws4.cell(row=row, column=2, value=gate["timing"]).border = thin_border
        ws4.cell(row=row, column=3, value=gate["criteria"]).border = thin_border
        ws4.cell(row=row, column=4, value=gate["owner"]).border = thin_border
        ws4.cell(row=row, column=5, value="").border = thin_border  # Status for user
        ws4.cell(row=row, column=6, value="").border = thin_border  # Actual date for user
        ws4.cell(row=row, column=7, value="").border = thin_border  # Notes for user
        for ci in range(1, len(gate_headers) + 1):
            ws4.cell(row=row, column=ci).alignment = wrap
        row += 1

    # Integration calendar
    row += 1
    ws4.cell(row=row, column=1, value="Financial Close Integration").font = Font(bold=True, size=11)
    row += 1
    cal_headers = ["Timing", "Financial Calendar Event", "SSBJ Action Required"]
    for ci, h in enumerate(cal_headers, 1):
        ws4.cell(row=row, column=ci, value=h)
    _apply_header(ws4, row, len(cal_headers))
    row += 1
    calendar_items = [
        ("Q1", "Q1 earnings / interim", "Begin GHG data collection for current FY. Establish data collection rhythm."),
        ("Q2", "Mid-year review", "Mid-year GHG data completeness check. Draft governance/risk disclosures."),
        ("Q3", "Q3 earnings", "9-month GHG data compiled. Internal dry run of SSBJ disclosures."),
        ("Q4", "Year-end close", "Finalize full-year GHG calculations. Complete all SSBJ disclosures."),
        ("Post-close", "Annual report filing (有報)", "Integrate SSBJ disclosure into securities report. Evidence freeze for assurance."),
    ]
    for timing, event, action in calendar_items:
        ws4.cell(row=row, column=1, value=timing).border = thin_border
        ws4.cell(row=row, column=2, value=event).border = thin_border
        ws4.cell(row=row, column=3, value=action).border = thin_border
        for ci in range(1, 4):
            ws4.cell(row=row, column=ci).alignment = wrap
        row += 1

    _auto_width(ws4)
    ws4.column_dimensions["C"].width = 55

    # ===== Sheet 5: Year 2+ Prep =====
    ws5 = wb.create_sheet("Year 2+ Preparation")
    ws5.sheet_properties.tabColor = "F97316"

    y2_headers = ["Criterion", "Category", "Year 1 Relief", "Year 1 Action Required",
                  "Year 2 Requirement", "Status", "Notes"]
    for ci, h in enumerate(y2_headers, 1):
        ws5.cell(row=1, column=ci, value=h)
    _apply_header(ws5, 1, len(y2_headers))

    row = 2
    for item in checklist_data["year2_prep"]:
        ws5.cell(row=row, column=1, value=item["criterion_id"]).border = thin_border
        ws5.cell(row=row, column=2, value=item["category"]).border = thin_border
        ws5.cell(row=row, column=3, value=item.get("relief_note", "")).border = thin_border
        ws5.cell(row=row, column=4, value=item.get("year1_action", "")).border = thin_border
        ws5.cell(row=row, column=5, value=item.get("year2_req", "")).border = thin_border
        ws5.cell(row=row, column=6, value="").border = thin_border  # Status for user
        ws5.cell(row=row, column=7, value="").border = thin_border  # Notes for user
        for ci in range(1, len(y2_headers) + 1):
            ws5.cell(row=row, column=ci).alignment = wrap
        row += 1

    # Assurance scope expansion
    row += 1
    ws5.cell(row=row, column=1, value="Assurance Scope Expansion Timeline").font = Font(bold=True, size=11)
    row += 1
    scope_headers = ["Year", "Assurance Scope", "What to Prepare"]
    for ci, h in enumerate(scope_headers, 1):
        ws5.cell(row=row, column=ci, value=h)
    _apply_header(ws5, row, len(scope_headers))
    row += 1
    scope_items = [
        ("Year 1", "Disclosure only (no assurance)", "Establish processes, collect evidence, engage assurance provider"),
        ("Year 2", "Limited assurance: Scope 1 & 2, Governance, Risk Mgmt", "Full evidence packages, controls operating effectively, auditor fieldwork"),
        ("Year 3+", "Scope expansion under consideration", "Progressive quantification, Scope 3 calculations, scenario analysis maturity"),
    ]
    for year, scope, prep in scope_items:
        ws5.cell(row=row, column=1, value=year).border = thin_border
        ws5.cell(row=row, column=2, value=scope).border = thin_border
        ws5.cell(row=row, column=3, value=prep).border = thin_border
        for ci in range(1, 4):
            ws5.cell(row=row, column=ci).alignment = wrap
        row += 1

    _auto_width(ws5)
    ws5.column_dimensions["C"].width = 50
    ws5.column_dimensions["D"].width = 50
    ws5.column_dimensions["E"].width = 50
    ws5.auto_filter.ref = f"A1:{get_column_letter(len(y2_headers))}{row - 1}"
    ws5.freeze_panes = "A2"

    # ===== Sheet 6: Review Findings (optional, only if review completed) =====
    if review_data:
        ws6 = wb.create_sheet("Review Findings")
        ws6.sheet_properties.tabColor = "10B981"

        # Header info
        ws6.merge_cells("A1:F1")
        ws6["A1"].value = f"Limited Assurance Review — {review_data.get('reviewer', '')} — {review_data.get('date', '')}"
        ws6["A1"].font = Font(size=12, bold=True)

        opinion = review_data.get("opinion", "")
        opinion_display = {"unqualified": "Unqualified (Clean)", "qualified": "Qualified",
                           "adverse": "Adverse", "disclaimer": "Disclaimer of Opinion"}.get(opinion, opinion)
        ws6.merge_cells("A2:F2")
        ws6["A2"].value = f"Overall Opinion: {opinion_display}"
        ws6["A2"].font = Font(size=11, bold=True,
                              color="16A34A" if opinion == "unqualified" else
                              "EAB308" if opinion == "qualified" else "DC2626")

        if review_data.get("findings"):
            ws6.merge_cells("A3:F3")
            ws6["A3"].value = f"Key Findings: {review_data['findings']}"
            ws6["A3"].alignment = wrap

        if review_data.get("recommendations"):
            ws6.merge_cells("A4:F4")
            ws6["A4"].value = f"Recommendations: {review_data['recommendations']}"
            ws6["A4"].alignment = wrap

        # Review items table
        rv_headers = ["Criterion", "Category", "Status", "Evidence Adequate", "Finding", "Recommendation"]
        start_row = 6
        for ci, h in enumerate(rv_headers, 1):
            ws6.cell(row=start_row, column=ci, value=h)
        _apply_header(ws6, start_row, len(rv_headers))

        row = start_row + 1
        for item in review_data.get("items", []):
            ws6.cell(row=row, column=1, value=item.get("criterion_id", "")).border = thin_border
            ws6.cell(row=row, column=2, value=item.get("category", "")).border = thin_border
            status_cell = ws6.cell(row=row, column=3, value=item.get("status", "").replace("_", " ").title())
            status_cell.border = thin_border
            if item.get("status") == "unsatisfactory":
                status_cell.font = Font(color="DC2626", bold=True)
            elif item.get("status") == "needs_improvement":
                status_cell.font = Font(color="EAB308", bold=True)
            elif item.get("status") == "satisfactory":
                status_cell.font = ok_font
            ev_cell = ws6.cell(row=row, column=4, value="Yes" if item.get("evidence_adequate") else "No")
            ev_cell.border = thin_border
            if not item.get("evidence_adequate"):
                ev_cell.font = Font(color="DC2626")
            ws6.cell(row=row, column=5, value=item.get("finding", "")).border = thin_border
            ws6.cell(row=row, column=6, value=item.get("recommendation", "")).border = thin_border
            for ci in range(1, len(rv_headers) + 1):
                ws6.cell(row=row, column=ci).alignment = wrap
            row += 1

        _auto_width(ws6)
        ws6.column_dimensions["E"].width = 50
        ws6.column_dimensions["F"].width = 50
        ws6.auto_filter.ref = f"A{start_row}:{get_column_letter(len(rv_headers))}{row - 1}"
        ws6.freeze_panes = f"A{start_row + 1}"

    return wb
